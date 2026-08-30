"""Import dormitory assignments from the university submission workbook."""

import re
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from dormitory.models import Dormitory, DormitoryAssignment
from generic.models import User


EXPECTED_HEADERS = ("学工号", "姓名", "住宿地址")
ADDRESS_PATTERN = re.compile(r"-(?P<room>\d+)-(?P<bed>\d+)号床\s*$")


class Command(BaseCommand):
    help = "Import dormitory assignments from a university submission workbook"

    def add_arguments(self, parser):
        parser.add_argument("excel_file", type=Path, help="Path to the Excel file")
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Validate the workbook without making database changes",
        )

    def handle(self, *args, **options):
        excel_file = options["excel_file"]
        assignments = self._read_assignments(excel_file)

        with transaction.atomic():
            for row_number, student_id, student_name, room_id, bed_id in assignments:
                try:
                    dormitory = Dormitory.objects.get(pk=room_id)
                except Dormitory.DoesNotExist as exc:
                    raise CommandError(
                        f"Row {row_number}: dormitory {room_id} does not exist"
                    ) from exc
                if not 1 <= bed_id <= dormitory.capacity:
                    raise CommandError(
                        f"Row {row_number}: bed {bed_id} is outside dormitory "
                        f"{room_id}'s capacity ({dormitory.capacity})"
                    )

                try:
                    user = User.objects.get(username=student_id)
                except User.DoesNotExist as exc:
                    self.stdout.write(
                        self.style.WARNING(
                            f"Row {row_number}: user {student_id} does not exist, skip"
                        )
                    )
                    continue
                if user.name != student_name:
                    raise CommandError(
                        f"Row {row_number}: name mismatch for {student_id}: "
                        f"workbook has {student_name!r}, database has {user.name!r}"
                    )

                _, created = DormitoryAssignment.objects.get_or_create(
                    dormitory=dormitory,
                    user=user,
                    bed_id=bed_id,
                )
                if not created:
                    self.stdout.write(
                        self.style.WARNING(
                            f"Assignment already exists: room {room_id}, "
                            f"user {student_id}, bed {bed_id}"
                        )
                    )

            if options["dry_run"]:
                transaction.set_rollback(True)
                self.stdout.write(
                    self.style.WARNING(
                        f"Dry run validated {len(assignments)} assignments; "
                        "no changes were made."
                    )
                )
            else:
                self.stdout.write(
                    self.style.SUCCESS(f"Imported {len(assignments)} assignments.")
                )

    def _read_assignments(self, path):
        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except (OSError, ValueError, KeyError) as exc:
            raise CommandError(f"Could not read workbook {path}: {exc}") from exc

        try:
            sheet = workbook.active
            headers = tuple(sheet.cell(1, column).value for column in (1, 2, 6))
            if headers != EXPECTED_HEADERS:
                raise CommandError(
                    f"Unexpected workbook format in {path}: columns A, B, and F "
                    f"must be {EXPECTED_HEADERS}, got {headers}"
                )

            assignments = []
            seen_students = set()
            seen_beds = set()
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=2, min_col=1, max_col=6, values_only=True),
                start=2,
            ):
                raw_student_id, student_name, *_, address = row
                if raw_student_id is None and student_name is None:
                    continue
                try:
                    student_id = str(int(raw_student_id))
                except (TypeError, ValueError) as exc:
                    raise CommandError(
                        f"Row {row_number}: invalid student ID {raw_student_id!r}"
                    ) from exc
                if not isinstance(student_name, str) or not student_name.strip():
                    raise CommandError(f"Row {row_number}: missing student name")
                match = ADDRESS_PATTERN.search(str(address))
                if match is None:
                    raise CommandError(
                        f"Row {row_number}: invalid accommodation address {address!r}"
                    )
                room_id = int(match.group("room"))
                bed_id = int(match.group("bed"))
                if student_id in seen_students:
                    raise CommandError(
                        f"Row {row_number}: duplicate student ID {student_id}"
                    )
                if (room_id, bed_id) in seen_beds:
                    raise CommandError(
                        f"Row {row_number}: duplicate bed {room_id}-{bed_id}"
                    )
                seen_students.add(student_id)
                seen_beds.add((room_id, bed_id))
                assignments.append(
                    (row_number, student_id, student_name.strip(), room_id, bed_id)
                )
            return assignments
        finally:
            workbook.close()
