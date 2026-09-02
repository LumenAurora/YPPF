"""Write dormitory assignments into the university-provided Excel template."""

import re
from argparse import ArgumentParser
from pathlib import Path
from typing import Any, TypeAlias

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# Sometimes some rooms that appear in the university-provided list can't be used.
# List them here.
RESTRICTED_ROOMS = {126, 164, 512, 547, 548, 556, 606, 668, 669}
ROOM_ID_PATTERN = re.compile(r"\d{3}")

StudentInfo: TypeAlias = dict[int, str]
RoomAssignments: TypeAlias = dict[int, list[tuple[int, str]]]


class Command(BaseCommand):
    help = (
        "Validate freshman and dormitory-assignment workbooks, write the "
        "assignments into a university template, and validate the result."
    )

    def add_arguments(self, parser: ArgumentParser) -> None:
        parser.add_argument(
            "--info",
            type=Path,
            default=Path("dormitory/references/info.xlsx"),
            help="Authoritative freshman workbook (default: dormitory/references/info.xlsx).",
        )
        parser.add_argument(
            "--assignments",
            type=Path,
            default=Path("dormitory/references/dorm_assigned.xlsx"),
            help="Dormitory assignment workbook (default: dormitory/references/dorm_assigned.xlsx).",
        )
        parser.add_argument(
            "--template",
            type=Path,
            default=Path("dormitory/references/template.xlsx"),
            help="University workbook template (default: dormitory/references/template.xlsx).",
        )
        parser.add_argument(
            "--output",
            type=Path,
            default=Path("dormitory/references/output.xlsx"),
            help="Destination workbook (default: dormitory/references/output.xlsx).",
        )

    def handle(self, *args: Any, **options: Any) -> None:
        info_path = options["info"]
        assignments_path = options["assignments"]
        template_path = options["template"]
        output_path = options["output"]

        student_info = self._read_student_info(info_path)
        self.stdout.write(
            f"There are {len(student_info)} students in the {info_path} file."
        )
        rooms, assigned_student_ids = self._read_assignments(
            assignments_path, student_info
        )

        self.stdout.write("Data validation passed. Now writing output workbook...")
        written_student_ids = self._write_output(template_path, output_path, rooms)
        difference = written_student_ids.symmetric_difference(assigned_student_ids)
        if difference:
            raise CommandError(
                f"Some students in {assignments_path} were not written to "
                f"{output_path}: {difference}"
            )

        self.stdout.write(f"Output done. Revalidating {output_path}...")
        self._validate_output(output_path, student_info)
        self.stdout.write(self.style.SUCCESS("Done!"))

    def _load_workbook(self, path: Path) -> Workbook:
        try:
            return openpyxl.load_workbook(path)
        except (OSError, ValueError, KeyError) as exc:
            raise CommandError(f"Could not read workbook {path}: {exc}") from exc

    def _sheet1(self, workbook: Workbook, path: Path) -> Worksheet:
        try:
            return workbook["Sheet1"]
        except KeyError as exc:
            raise CommandError(f"Workbook {path} has no 'Sheet1' worksheet") from exc

    def _read_student_info(self, path: Path) -> StudentInfo:
        workbook = self._load_workbook(path)
        try:
            sheet = self._sheet1(workbook, path)
            student_info: StudentInfo = {}
            for sid, name in sheet.iter_rows(
                min_row=2, min_col=1, max_col=2, values_only=True
            ):
                try:
                    sid = int(sid)
                except (TypeError, ValueError) as exc:
                    raise CommandError(
                        f"Invalid student ID {sid!r} in {path}"
                    ) from exc
                if sid in student_info:
                    raise CommandError(f"Duplicate student ID found: {sid}")
                student_info[sid] = name
            return student_info
        finally:
            workbook.close()

    def _read_assignments(
        self, path: Path, student_info: StudentInfo
    ) -> tuple[RoomAssignments, set[int]]:
        workbook = self._load_workbook(path)
        try:
            sheet = self._sheet1(workbook, path)
            rooms: RoomAssignments = {}
            seen_student_ids: set[int] = set()
            for rid, name, gender, sid in sheet.iter_rows(
                min_row=2, min_col=1, max_col=4, values_only=True
            ):
                try:
                    rid, sid = int(rid), int(sid)
                except (TypeError, ValueError) as exc:
                    raise CommandError(
                        f"Invalid room or student ID in {path}: {rid!r}, {sid!r}"
                    ) from exc
                if sid not in student_info:
                    raise CommandError(
                        f"Student ID {sid} in {path} not found in authoritative data"
                    )
                if student_info[sid] != name:
                    raise CommandError(
                        f"Name mismatch for student ID {sid}: authoritative data has "
                        f"'{student_info[sid]}', {path} has '{name}'"
                    )
                if rid in RESTRICTED_ROOMS:
                    raise CommandError(
                        f"Student ID {sid} is assigned to a restricted room {rid}"
                    )
                if gender == "男" and rid > 500:
                    self.stderr.write(
                        f"Male student ID {sid} is assigned to a female dorm room {rid}"
                    )
                if gender == "女" and rid < 500:
                    self.stderr.write(
                        f"Female student ID {sid} is assigned to a male dorm room {rid}"
                    )
                if sid in seen_student_ids:
                    raise CommandError(f"Duplicate student ID {sid} found in {path}")
                seen_student_ids.add(sid)
                rooms.setdefault(rid, []).append((sid, name))
                if len(rooms[rid]) > 4:
                    raise CommandError(f"Room {rid} has more than 4 students assigned")

            missing = set(student_info) - seen_student_ids
            if missing:
                raise CommandError(
                    "Some students in the authoritative workbook are not assigned "
                    f"to any dorm room: {missing}"
                )
            return rooms, seen_student_ids
        finally:
            workbook.close()

    def _write_output(
        self,
        template_path: Path,
        output_path: Path,
        rooms: RoomAssignments,
    ) -> set[int]:
        workbook = self._load_workbook(template_path)
        written_student_ids: set[int] = set()
        try:
            for row in workbook.active.iter_rows(min_row=2):
                match = ROOM_ID_PATTERN.search(str(row[5].value))
                if match is None:
                    self.stderr.write(
                        f"Warning: No room ID found in row {row[5].row}, column 6"
                    )
                    continue
                rid = int(match.group(0))
                if rooms.get(rid):
                    sid, name = rooms[rid].pop(0)
                    row[0].value = sid
                    row[1].value = name
                    written_student_ids.add(sid)
            try:
                workbook.save(output_path)
            except OSError as exc:
                raise CommandError(
                    f"Could not write output workbook {output_path}: {exc}"
                ) from exc
            return written_student_ids
        finally:
            workbook.close()

    def _validate_output(
        self, path: Path, expected_student_info: StudentInfo
    ) -> None:
        workbook = self._load_workbook(path)
        try:
            output_student_info: StudentInfo = {}
            for sid, name in workbook.active.iter_rows(
                min_row=2, min_col=1, max_col=2, values_only=True
            ):
                if (sid is None) != (name is None):
                    raise CommandError(
                        f"A row in {path} has only one of student ID or name filled"
                    )
                if sid is None:
                    continue
                try:
                    sid = int(sid)
                except (TypeError, ValueError) as exc:
                    raise CommandError(f"Invalid student ID {sid!r} in {path}") from exc
                if sid in output_student_info:
                    raise CommandError(f"Duplicate student ID {sid} found in {path}")
                output_student_info[sid] = name

            if output_student_info != expected_student_info:
                difference = set(output_student_info.items()).symmetric_difference(
                    set(expected_student_info.items())
                )
                raise CommandError(
                    f"Mismatch between authoritative data and {path}: {difference}"
                )
        finally:
            workbook.close()
