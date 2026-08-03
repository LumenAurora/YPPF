import os
import tempfile
from datetime import timedelta

from django.core.management import call_command
from django.core.exceptions import ValidationError
from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook

from generic.models import User
from questionnaire.models import AnswerSheet, AnswerText, Choice, Question, Survey
from questionnaire.serializers import AnswerTextSerializer
from questionnaire.validators import validate_answer_body
from questionnaire.management.commands.dump_questionnaire_result import Command


class RankingQuestionTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user(
            username='tester',
            name='tester',
            password='password',
        )
        now = timezone.now()
        self.survey = Survey.objects.create(
            title='Ranking Survey',
            description='test',
            creator=self.owner,
            status=Survey.Status.PUBLISHED,
            start_time=now,
            end_time=now + timedelta(days=1),
        )
        self.question = Question.objects.create(
            survey=self.survey,
            order=1,
            topic='Rank these options',
            type=Question.Type.RANKING,
            required=True,
        )
        self.choice1 = Choice.objects.create(
            question=self.question,
            order=1,
            text='Alpha',
        )
        self.choice2 = Choice.objects.create(
            question=self.question,
            order=2,
            text='Beta',
        )
        self.choice3 = Choice.objects.create(
            question=self.question,
            order=3,
            text='Gamma',
        )
        self.sheet = AnswerSheet.objects.create(
            survey=self.survey,
            creator=self.owner,
        )

    def test_ranking_serializer_requires_all_distinct_choices(self):
        serializer = AnswerTextSerializer(data={
            'question': self.question.id,
            'answersheet': self.sheet.id,
            'body': '2,2,1',
        })
        self.assertFalse(serializer.is_valid())
        self.assertIn('排序题不允许重复选项', str(serializer.errors))

    def test_export_command_outputs_ranking_text(self):
        AnswerText.objects.create(
            question=self.question,
            answersheet=self.sheet,
            body='2,1,3',
        )

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_file = os.path.join(tmp_dir, 'result.xlsx')
            call_command('dump_questionnaire_result', self.survey.title, output_file)

            wb = load_workbook(output_file)
            ws = wb.active
            values = [
                ws.cell(row=row_index, column=1).value
                for row_index in range(2, ws.max_row + 1)
            ]
            self.assertIn('1. Beta; 2. Alpha; 3. Gamma', values)

    def test_decode_choice_text_uses_in_memory_map(self):
        choice_text_by_key = {
            (self.question.id, 1): 'Alpha',
            (self.question.id, 2): 'Beta',
            (self.question.id, 3): 'Gamma',
        }

        with self.assertNumQueries(0):
            texts = Command._decode_choice_text(
                self.question, '2, 1,3', choice_text_by_key,
            )

        self.assertEqual(texts, ['Beta', 'Alpha', 'Gamma'])


class AnswerBodyValidationTests(TestCase):
    def setUp(self):
        owner = User.objects.create_user(username='validator', name='Validator')
        now = timezone.now()
        survey = Survey.objects.create(
            title='Validation Survey',
            creator=owner,
            start_time=now,
            end_time=now + timedelta(days=1),
        )
        self.question = Question.objects.create(
            survey=survey,
            order=1,
            topic='Choose',
            type=Question.Type.MULTIPLE,
            min_choices=2,
            max_choices=3,
        )
        for order in range(1, 5):
            Choice.objects.create(
                question=self.question, order=order, text=str(order))

    def test_rejects_choice_outside_question_range(self):
        with self.assertRaisesMessage(ValidationError, '超出有效范围'):
            validate_answer_body(self.question, '1,5')

    def test_rejects_duplicate_choices(self):
        with self.assertRaisesMessage(ValidationError, '不允许重复选项'):
            validate_answer_body(self.question, '1,1')

    def test_enforces_minimum_and_maximum_choices(self):
        with self.assertRaisesMessage(ValidationError, '至少需要选择 2 个'):
            validate_answer_body(self.question, '1')
        with self.assertRaisesMessage(ValidationError, '最多只能选择 3 个'):
            validate_answer_body(self.question, '1,2,3,4')

    def test_ranking_requires_every_choice_exactly_once(self):
        self.question.type = Question.Type.RANKING
        with self.assertRaisesMessage(ValidationError, '需要包含所有选项'):
            validate_answer_body(self.question, '1,2,3')
        with self.assertRaisesMessage(ValidationError, '不允许重复选项'):
            validate_answer_body(self.question, '1,2,3,3')
