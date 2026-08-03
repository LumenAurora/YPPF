from openpyxl import Workbook
from django.core.management.base import BaseCommand, CommandParser
from tqdm import tqdm

from questionnaire.models import Survey, AnswerSheet, Choice
import os


class Command(BaseCommand):
    help = 'Dumps the result of a questionnaire to raw_data/result.xlsx'

    @staticmethod
    def _decode_choice_text(question, body, choice_text_by_key):
        choice_orders = [segment.strip() for segment in body.split(',') if segment.strip()]
        try:
            return [
                choice_text_by_key[(question.id, int(order))]
                for order in choice_orders
            ]
        except KeyError as error:
            raise ValueError(
                f'Choice {error.args[0][1]} does not exist for question {question.id}'
            ) from error

    def add_arguments(self, parser: CommandParser) -> None:
        parser.add_argument('questionnaire_title', type=str,
                            help='Title of questionnaire to dump')
        parser.add_argument('output_file', type=str, default='raw_data/result.xlsx',
                            help='Output file path')
        return super().add_arguments(parser)

    def handle(self, *args, **options):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Result'

        survey = Survey.objects.get(title=options['questionnaire_title'])
        questions = survey.questions.order_by('order').all()
        choice_text_by_key = {
            (question_id, order): text
            for question_id, order, text in Choice.objects.filter(
                question__survey=survey,
            ).values_list('question_id', 'order', 'text')
        }

        # 若没有该文件，自动创建
        if not os.path.exists(options['output_file']):
            os.makedirs(os.path.dirname(options['output_file']), exist_ok=True)
        self.stdout.write(self.style.NOTICE(f'Survey title: {survey.title}'))

        # Add header row
        headers = [question.topic for question in questions]
        for col_num, column_title in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=column_title)

        answer_sheets = AnswerSheet.objects.filter(survey=survey)

        # Iterate through the AnswerSheet objects
        for row_num, answer_sheet in tqdm(
            enumerate(answer_sheets, 2),
            total=answer_sheets.count(),
            desc='Processing answer sheets',
        ):
            answers = {
                answer.question.id: answer
                for answer in answer_sheet.answertext_set.all()
            }
            for col_num, question in enumerate(questions, 1):
                answer = answers.get(question.id)
                value = None
                if answer:
                    if question.type == 'TEXT':
                        value = answer.body
                    elif question.type == 'SINGLE':
                        value = self._decode_choice_text(
                            question, answer.body, choice_text_by_key,
                        )[0]
                    elif question.type == 'MULTIPLE':
                        choices_texts = self._decode_choice_text(
                            question, answer.body, choice_text_by_key,
                        )
                        value = ', '.join(choices_texts)
                    elif question.type == 'RANKING':
                        choices_texts = self._decode_choice_text(
                            question, answer.body, choice_text_by_key,
                        )
                        ranking_lines = [
                            f'{rank}. {text}'
                            for rank, text in enumerate(choices_texts, start=1)
                        ]
                        value = '; '.join(ranking_lines)
                ws.cell(row=row_num, column=col_num, value=value)

        # Save and return the workbook
        with open(options['output_file'], 'wb') as f:
            wb.save(f)

        self.stdout.write(self.style.SUCCESS(
            f'Results dumped to {options["output_file"]}'
        ))
